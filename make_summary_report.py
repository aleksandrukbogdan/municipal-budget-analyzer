"""
Формирование сводной таблицы бюджетов МО.

Описание:
    Создаёт форматированный Excel-отчёт с данными по капитальным
    расходам муниципальных образований. Поддерживает динамическое
    количество годов и МО.

Структура отчёта:
    - Отдельный лист для каждого года (2024, 2025, 2026, ...)
    - Лист "Динамика" - сравнение по годам с расчётом изменений
    - Строки: МО (муниципальные образования)
    - Столбцы:
        * Доходы: Всего, Безвозмездные, Налоговые (заглушки)
        * Расходы: Всего + 4 категории (абс. значения и %)
        * Дефицит/профицит (заглушка)

Категории расходов:
    - Транспортная инфраструктура
    - Благоустройство и ЖКХ
    - Объекты образования
    - Объекты здравоохранения, культуры, спорта

Входные данные:
    - output/budget_capital_expenses.xlsx - данные из calculate_costs.py

Выходные данные:
    - output/Сводный_Отчет_Бюджет.xlsx - форматированный отчёт

Использование:
    python make_summary_report.py [input.xlsx]

    или как модуль:
    from make_summary_report import create_summary_report
    create_summary_report(input_file, output_file)

Автор: Автоматически сгенерировано
Версия: 1.0
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# Категории расходов (из нашей классификации)
EXPENSE_CATEGORIES = [
    "Транспортная инфраструктура",
    "Благоустройство и ЖКХ",
    "Объекты образования",
    "Объекты здравоохранения, культуры, спорта"
]

# Стили
HEADER_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
HEADER_FILL_YELLOW = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def load_budget_data(input_file=None):
    """Загружает данные о расходах из Excel."""
    if input_file is None:
        if os.path.exists("output/budget_capital_expenses.xlsx"):
            input_file = "output/budget_capital_expenses.xlsx"
        else:
            input_file = "output/extracted_budget_data.xlsx"
    
    if not os.path.exists(input_file):
        print(f"Файл {input_file} не найден!")
        return None
    
    df = pd.read_excel(input_file)
    return df


def prepare_summary_data(df):
    """
    Подготавливает сводные данные по МО и категориям.
    Года определяются динамически из данных.
    Все МО присутствуют на всех листах (с пустыми значениями если данных нет).
    
    Returns:
        dict: {year: DataFrame с данными по МО}
        list: список всех годов
    """
    # Определяем колонки с годами (динамически!)
    year_cols = []
    for c in df.columns:
        try:
            if isinstance(c, (int, float)):
                y = int(c)
            else:
                y = int(str(c))
            if 2020 <= y <= 2100:
                year_cols.append(y)
        except:
            pass
    
    year_cols = sorted(set(year_cols))
    
    if not year_cols:
        print("Года не найдены в данных!")
        return {}, []
    
    print(f"Найдены года (динамически): {year_cols}")
    
    # Собираем ВСЕ уникальные МО из данных
    all_mo = sorted(df['МО'].unique().tolist())
    print(f"Найдено МО: {len(all_mo)}")
    
    result = {}
    
    for year in year_cols:
        # Определяем имя колонки (может быть int или str)
        year_col = year if year in df.columns else str(year)
        
        if year_col not in df.columns:
            print(f"  Предупреждение: колонка {year} не найдена")
            continue
        
        # Группируем по МО и категории
        grouped = df.groupby(['МО', 'Категория'])[year_col].sum().reset_index()
        
        # Преобразуем в широкий формат (pivot)
        pivot = grouped.pivot(index='МО', columns='Категория', values=year_col)
        
        # Добавляем ВСЕ МО (даже если у них нет данных за этот год)
        pivot = pivot.reindex(all_mo)
        
        # Заполняем пропуски нулями
        pivot = pivot.fillna(0)
        
        # Убеждаемся что все категории есть
        for cat in EXPENSE_CATEGORIES:
            if cat not in pivot.columns:
                pivot[cat] = 0.0
        
        # Упорядочиваем категории
        pivot = pivot[EXPENSE_CATEGORIES]
        
        # Добавляем итого расходов
        pivot['Расходы_Всего'] = pivot[EXPENSE_CATEGORIES].sum(axis=1)
        
        # Рассчитываем проценты (защита от деления на 0)
        for cat in EXPENSE_CATEGORIES:
            pct_col = f'{cat}_%'
            pivot[pct_col] = pivot.apply(
                lambda row: (row[cat] / row['Расходы_Всего'] * 100) if row['Расходы_Всего'] > 0 else 0,
                axis=1
            )
        
        result[year] = pivot.reset_index()
        print(f"  {year}: {len(pivot)} МО, сумма: {pivot['Расходы_Всего'].sum():,.0f}")
    
    return result, year_cols


def create_year_sheet(ws, data_df, year):
    """
    Создаёт лист для конкретного года.
    
    Структура столбцов:
    № | МО | Население | Бюдж.обесп.(Доходы) | Бюдж.обесп.(Расходы) |
    Доходы: Всего | Безвозм. | % | Налог. | % |
    Расходы: Всего | Трансп. | % | ЖКХ | % | Образ. | % | Здрав. | % |
    Дефицит
    """
    
    # === ЗАГОЛОВКИ (строки 1-5) ===
    
    # Строка 1-2: Основные группы
    headers_row1 = [
        ('№ п/п', 1, 5, 1, 1),
        ('Муниципальное образование', 2, 5, 2, 2),
        ('Население*', 3, 5, 3, 3),
        ('Бюджетная обеспеченность на человека', 4, 4, 4, 5),
        ('Доходы', 6, 3, 6, 10),
        ('Расходы', 11, 2, 11, 19),
        ('Дефицит/ профицит бюджета', 20, 5, 20, 20),
    ]
    
    # Строка 3: Подгруппы бюджетной обеспеченности
    # Строка 4: Доходы и расходы подгруппы
    headers_row4_income = ['Всего', 'Безвозмездные поступления', '', 'Налоговые и неналоговые доходы', '']
    headers_row4_expense = ['Всего', 'Транспортная инфраструктура', '', 'Благоустройство и ЖКХ', '', 
                           'Объекты образования', '', 'Объекты здравоохранения, культуры, спорта', '']
    
    # Строка 5: тыс.руб. и %
    
    # Применяем заголовки
    ws.merge_cells('A1:A5')
    ws['A1'] = '№ п/п'
    
    ws.merge_cells('B1:B5')
    ws['B1'] = 'Муниципальное образование'
    
    ws.merge_cells('C1:C5')
    ws['C1'] = 'Население*'
    
    ws.merge_cells('D1:E2')
    ws['D1'] = 'Бюджетная обеспеченность на человека'
    ws['D3'] = 'Доходы'
    ws['E3'] = 'Расходы'
    ws['D4'] = 'тыс. руб.'
    ws['E4'] = 'тыс. руб.'
    ws.merge_cells('D4:D5')
    ws.merge_cells('E4:E5')
    
    # Доходы
    ws.merge_cells('F1:J2')
    ws['F1'] = 'Доходы'
    ws.merge_cells('F3:F5')
    ws['F3'] = 'Всего'
    ws.merge_cells('G3:H3')
    ws['G3'] = 'в том числе'
    ws.merge_cells('G4:H4')
    ws['G4'] = 'Безвозмездные поступления'
    ws['G5'] = 'тыс. руб.'
    ws['H5'] = '%'
    ws.merge_cells('I4:J4')
    ws['I4'] = 'Налоговые и неналоговые доходы'
    ws['I5'] = 'тыс. руб.'
    ws['J5'] = '%'
    
    # Расходы
    ws.merge_cells('K1:S1')
    ws['K1'] = 'Расходы'
    ws.merge_cells('K2:K5')
    ws['K2'] = 'Всего'
    ws.merge_cells('L2:S2')
    ws['L2'] = 'в том числе расходы на социальные объекты'
    
    # Категории расходов
    ws.merge_cells('L3:M3')
    ws['L3'] = 'Транспортная инфраструктура'
    ws['L4'] = 'тыс. руб.'
    ws['M4'] = '%'
    ws.merge_cells('L4:L5')
    ws.merge_cells('M4:M5')
    
    ws.merge_cells('N3:O3')
    ws['N3'] = 'Благоустройство и ЖКХ'
    ws['N4'] = 'тыс. руб.'
    ws['O4'] = '%'
    ws.merge_cells('N4:N5')
    ws.merge_cells('O4:O5')
    
    ws.merge_cells('P3:Q3')
    ws['P3'] = 'Объекты образования'
    ws['P4'] = 'тыс. руб.'
    ws['Q4'] = '%'
    ws.merge_cells('P4:P5')
    ws.merge_cells('Q4:Q5')
    
    ws.merge_cells('R3:S3')
    ws['R3'] = 'Объекты здравоохранения, культуры, спорта'
    ws['R4'] = 'тыс. руб.'
    ws['S4'] = '%'
    ws.merge_cells('R4:R5')
    ws.merge_cells('S4:S5')
    
    # Дефицит
    ws.merge_cells('T1:T5')
    ws['T1'] = 'Дефицит/ профицит бюджета'
    
    # === ФОРМАТИРОВАНИЕ ЗАГОЛОВКОВ ===
    for row in range(1, 6):
        for col in range(1, 21):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
            if col <= 5:
                cell.fill = HEADER_FILL
            elif col <= 10:
                cell.fill = HEADER_FILL
            else:
                cell.fill = HEADER_FILL_GREEN
    
    # === ДАННЫЕ ===
    start_row = 6
    
    for idx, row_data in data_df.iterrows():
        row_num = start_row + idx
        mo_name = row_data['МО']
        
        # № п/п
        ws.cell(row=row_num, column=1, value=idx + 1)
        
        # МО
        ws.cell(row=row_num, column=2, value=mo_name)
        
        # Население (заглушка)
        ws.cell(row=row_num, column=3, value=0)
        
        # Бюджетная обеспеченность (заглушки)
        ws.cell(row=row_num, column=4, value=0)  # Доходы на чел.
        ws.cell(row=row_num, column=5, value=0)  # Расходы на чел.
        
        # Доходы (заглушки)
        ws.cell(row=row_num, column=6, value=0)  # Всего
        ws.cell(row=row_num, column=7, value=0)  # Безвозмездные
        ws.cell(row=row_num, column=8, value=0)  # %
        ws.cell(row=row_num, column=9, value=0)  # Налоговые
        ws.cell(row=row_num, column=10, value=0)  # %
        
        # Расходы - данные из нашей таблицы
        total_expenses = row_data.get('Расходы_Всего', 0)
        ws.cell(row=row_num, column=11, value=total_expenses)
        
        # Транспортная инфраструктура
        val = row_data.get('Транспортная инфраструктура', 0)
        pct = row_data.get('Транспортная инфраструктура_%', 0)
        ws.cell(row=row_num, column=12, value=val)
        ws.cell(row=row_num, column=13, value=pct)
        
        # Благоустройство и ЖКХ
        val = row_data.get('Благоустройство и ЖКХ', 0)
        pct = row_data.get('Благоустройство и ЖКХ_%', 0)
        ws.cell(row=row_num, column=14, value=val)
        ws.cell(row=row_num, column=15, value=pct)
        
        # Объекты образования
        val = row_data.get('Объекты образования', 0)
        pct = row_data.get('Объекты образования_%', 0)
        ws.cell(row=row_num, column=16, value=val)
        ws.cell(row=row_num, column=17, value=pct)
        
        # Объекты здравоохранения
        val = row_data.get('Объекты здравоохранения, культуры, спорта', 0)
        pct = row_data.get('Объекты здравоохранения, культуры, спорта_%', 0)
        ws.cell(row=row_num, column=18, value=val)
        ws.cell(row=row_num, column=19, value=pct)
        
        # Дефицит/профицит (заглушка)
        ws.cell(row=row_num, column=20, value=0)
        
        # Форматирование строки данных
        for col in range(1, 21):
            cell = ws.cell(row=row_num, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right' if col > 2 else 'left', vertical='center')
            
            # Форматирование чисел
            if col > 2:
                if col in [8, 10, 13, 15, 17, 19]:  # Проценты
                    cell.number_format = '0.0%' if isinstance(cell.value, (int, float)) and cell.value < 1 else '0.0'
                else:
                    cell.number_format = '#,##0.00'
    
    # === СТРОКА ИТОГО ===
    total_row = start_row + len(data_df)
    ws.cell(row=total_row, column=2, value='ИТОГО')
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    
    # Суммируем по колонкам
    for col in [3, 6, 7, 9, 11, 12, 14, 16, 18, 20]:
        total = sum(ws.cell(row=r, column=col).value or 0 for r in range(start_row, total_row))
        ws.cell(row=total_row, column=col, value=total)
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).number_format = '#,##0.00'
    
    for col in range(1, 21):
        ws.cell(row=total_row, column=col).border = THIN_BORDER
    
    # === ШИРИНА СТОЛБЦОВ ===
    column_widths = {
        'A': 6, 'B': 25, 'C': 12, 'D': 12, 'E': 12,
        'F': 15, 'G': 15, 'H': 8, 'I': 15, 'J': 8,
        'K': 15, 'L': 15, 'M': 8, 'N': 15, 'O': 8,
        'P': 15, 'Q': 8, 'R': 15, 'S': 8, 'T': 15
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Высота заголовков
    for row in range(1, 6):
        ws.row_dimensions[row].height = 30


def create_dynamics_sheet(ws, all_data, years):
    """Создаёт лист с динамикой по годам (динамическое количество годов)."""
    
    # Заголовок с динамической шириной
    last_col = len(years) + 3  # МО + года + изменение + % изменения
    ws['A1'] = 'Динамика расходов на капитальные затраты по годам'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A1:{get_column_letter(last_col)}1')
    
    # Собираем все МО
    all_mo = set()
    for year, df in all_data.items():
        all_mo.update(df['МО'].tolist())
    all_mo = sorted(list(all_mo))
    
    # Заголовки
    row = 3
    ws.cell(row=row, column=1, value='№')
    ws.cell(row=row, column=2, value='Муниципальное образование')
    col = 3
    for year in years:
        ws.cell(row=row, column=col, value=f'{year}')
        col += 1
    ws.cell(row=row, column=col, value='Изменение')
    ws.cell(row=row, column=col + 1, value='% изм.')
    
    # Подзаголовок (единицы измерения)
    row = 4
    ws.cell(row=row, column=1, value='')
    ws.cell(row=row, column=2, value='')
    col = 3
    for year in years:
        ws.cell(row=row, column=col, value='тыс. руб.')
        col += 1
    ws.cell(row=row, column=col, value='тыс. руб.')
    ws.cell(row=row, column=col + 1, value='%')
    
    # Форматируем заголовки
    for r in [3, 4]:
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(bold=True, size=10)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Объединяем ячейки заголовков
    ws.merge_cells('A3:A4')
    ws.merge_cells('B3:B4')
    
    # Данные по МО
    data_row = 5
    for idx, mo in enumerate(all_mo):
        ws.cell(row=data_row, column=1, value=idx + 1)
        ws.cell(row=data_row, column=1).border = THIN_BORDER
        ws.cell(row=data_row, column=1).alignment = Alignment(horizontal='center')
        
        ws.cell(row=data_row, column=2, value=mo)
        ws.cell(row=data_row, column=2).border = THIN_BORDER
        
        values = []
        col = 3
        for year in years:
            if year in all_data:
                df = all_data[year]
                mo_data = df[df['МО'] == mo]
                if not mo_data.empty:
                    val = mo_data['Расходы_Всего'].values[0]
                else:
                    val = 0
            else:
                val = 0
            
            values.append(val)
            cell = ws.cell(row=data_row, column=col, value=val if val > 0 else '')
            cell.number_format = '#,##0.00'
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='right')
            col += 1
        
        # Изменение (последний год - первый год)
        first_val = next((v for v in values if v > 0), 0)
        last_val = next((v for v in reversed(values) if v > 0), 0)
        
        if first_val > 0 and last_val > 0:
            change = last_val - first_val
            pct_change = (change / first_val * 100) if first_val > 0 else 0
        else:
            change = ''
            pct_change = ''
        
        cell = ws.cell(row=data_row, column=col, value=change)
        cell.number_format = '#,##0.00'
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
        
        cell = ws.cell(row=data_row, column=col + 1, value=pct_change)
        if isinstance(pct_change, (int, float)):
            cell.number_format = '0.0'
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='right')
        
        data_row += 1
    
    # Строка ИТОГО
    ws.cell(row=data_row, column=2, value='ИТОГО')
    ws.cell(row=data_row, column=2).font = Font(bold=True)
    ws.cell(row=data_row, column=2).border = THIN_BORDER
    
    col = 3
    totals = []
    for year in years:
        total = sum(
            ws.cell(row=r, column=col).value or 0 
            for r in range(5, data_row)
        )
        totals.append(total)
        cell = ws.cell(row=data_row, column=col, value=total)
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        col += 1
    
    # Итого изменение
    if len(totals) >= 2 and totals[0] > 0:
        total_change = totals[-1] - totals[0]
        total_pct = (total_change / totals[0] * 100)
    else:
        total_change = 0
        total_pct = 0
    
    cell = ws.cell(row=data_row, column=col, value=total_change)
    cell.number_format = '#,##0.00'
    cell.font = Font(bold=True)
    cell.border = THIN_BORDER
    
    cell = ws.cell(row=data_row, column=col + 1, value=total_pct)
    cell.number_format = '0.0'
    cell.font = Font(bold=True)
    cell.border = THIN_BORDER
    
    # Ширина столбцов
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    for i in range(len(years)):
        ws.column_dimensions[get_column_letter(i + 3)].width = 16
    ws.column_dimensions[get_column_letter(len(years) + 3)].width = 14
    ws.column_dimensions[get_column_letter(len(years) + 4)].width = 10


def create_summary_report(input_file=None, output_file=None):
    """
    Создаёт сводный отчёт с несколькими листами.
    """
    if output_file is None:
        output_file = "output/Сводный_Отчет_Бюджет.xlsx"
    
    # Загружаем данные
    df = load_budget_data(input_file)
    if df is None:
        return
    
    print(f"Загружено строк: {len(df)}")
    
    # Подготавливаем сводные данные
    all_data, years = prepare_summary_data(df)
    
    if not all_data:
        print("Нет данных для отчёта!")
        return
    
    print(f"Подготовлены данные по годам: {years}")
    
    # Создаём workbook
    wb = openpyxl.Workbook()
    
    # Удаляем дефолтный лист
    default_sheet = wb.active
    
    # Создаём лист для каждого года
    for year in years:
        if year in all_data:
            ws = wb.create_sheet(title=f'{year}')
            create_year_sheet(ws, all_data[year], year)
            print(f"  Создан лист: {year}")
    
    # Создаём лист динамики
    ws_dynamics = wb.create_sheet(title='Динамика')
    create_dynamics_sheet(ws_dynamics, all_data, years)
    print("  Создан лист: Динамика")
    
    # Удаляем дефолтный лист
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Сохраняем
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    wb.save(output_file)
    
    print(f"\nСводный отчёт создан: {output_file}")
    
    # Выводим сводку
    print("\nСводка по данным:")
    for year in years:
        if year in all_data:
            total = all_data[year]['Расходы_Всего'].sum()
            mo_count = len(all_data[year])
            print(f"  {year}: {mo_count} МО, итого расходов: {total:,.2f} тыс.руб.")


if __name__ == "__main__":
    import sys
    
    input_file = sys.argv[1] if len(sys.argv) > 1 else None
    create_summary_report(input_file)
